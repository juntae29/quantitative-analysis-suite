import pandas as pd
import pypdf
import io
import re
from openpyxl import load_workbook # 안전한 엑셀 로딩을 위해 추가

class DataProcessor:
    def load_file(self, file):
        if file.name.lower().endswith(('.ttf', '.py', '.js', '.css', '.yaml', '.txt')):
            return None
            
        content = file.getvalue()
        try:
            # 1. PDF 파일 처리
            if file.name.lower().endswith('.pdf'):
                reader = pypdf.PdfReader(io.BytesIO(content))
                text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
                df = pd.DataFrame({'combined': [text]})
                return df[['combined']]
                
            # 2. Excel (XLSX, XLS) 파일 처리 (최신 파이썬 버전 호환성 완벽 우회)
            elif file.name.lower().endswith(('.xlsx', '.xls')):
                # BytesIO 호환성 에러를 피하기 위해 openpyxl로 수동 파싱 진행
                wb = load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                
                data = []
                for row in ws.iter_rows(values_only=True):
                    # 모든 셀이 비어있지 않은 행만 수집
                    if any(cell is not None for cell in row):
                        data.append([str(cell) if cell is not None else "" for cell in row])
                
                if not data:
                    return None
                    
                # 데이터프레임 빌드
                df = pd.DataFrame(data[1:], columns=data[0]) if len(data) > 1 else pd.DataFrame(data)
                
            # 3. CSV 파일 처리
            else:
                df = pd.read_csv(io.BytesIO(content), encoding='utf-8-sig')
            
            # CSV 및 Excel 파일 공통 데이터 정제 및 병합
            df = df.fillna('').astype(str)
            if not df.empty:
                df['combined'] = df.apply(lambda row: ' '.join(row.values), axis=1)
                return df[['combined']]
            else:
                return None
                
        except Exception:
            return None

    def normalize(self, text):
        if not isinstance(text, str):
            text = str(text) if text is not None else ""
        
        text = re.sub(r'[^가-힣a-zA-Z0-9\s]', ' ', text).lower()
        words = text.split()
        
        noise_words = {
            'fonts', 'visualizer', 'process', 'engine', 'main', 'py', 'ttf', 'otf', 
            'csv', 'xlsx', 'pdf', 'modules', 'data', 'git', 'commit', 'push', 
            'run', 'fix', 'cloud', 'bash', 'filter', 'rerun', 'loading', 'st', 'streamlit'
        }
        
        return [w for w in words if (len(w) > 1 or re.match(r'[가-힣]', w)) and w not in noise_words]